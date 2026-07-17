"""Import học sinh từ Excel: parse -> validate từng dòng -> preview -> commit.
Cột: họ tên*, ngày sinh*, giới tính, email, SĐT phụ huynh, mã lớp, ghi chú.
Xem SRS ORG §5.2, FR-ORG-09/10/11.
"""

import io
import json
import uuid
from datetime import date, datetime

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.org import users_service as us

MAX_ROWS = 2000
_HEADERS = ["full_name", "dob", "gender", "email", "parent_phone", "class_name", "note"]


class TooManyRows(Exception):
    pass


def _parse_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError("bad_date")


async def _class_map(s: AsyncSession) -> dict[str, str]:
    rows = (await s.execute(text("SELECT id, name FROM classes"))).mappings().all()
    return {r["name"]: str(r["id"]) for r in rows}


async def validate_file(
    s: AsyncSession, tenant_id: str, filename: str, content: bytes, created_by: str
) -> dict:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    if len(data_rows) > MAX_ROWS:
        raise TooManyRows

    class_map = await _class_map(s)
    seen_names: set[tuple] = set()
    parsed: list[dict] = []
    valid = errors = 0

    for idx, raw in enumerate(data_rows, start=2):
        cells = list(raw) + [None] * (len(_HEADERS) - len(raw))
        rec = dict(zip(_HEADERS, cells, strict=False))
        err = None
        dob = None
        if not rec.get("full_name"):
            err = "Thiếu họ tên"
        else:
            try:
                dob = _parse_date(rec.get("dob"))
            except ValueError:
                err = "Sai định dạng ngày sinh"
            if err is None:
                cn = rec.get("class_name")
                if cn and cn not in class_map:
                    err = f"Mã lớp '{cn}' không tồn tại"
            if err is None:
                key = (str(rec["full_name"]).strip().lower(), str(dob))
                if key in seen_names:
                    err = "Trùng trong file"
                seen_names.add(key)
        rec["dob"] = str(dob) if dob else None
        parsed.append(
            {
                "row_no": idx,
                "data": {k: (str(rec[k]) if rec[k] is not None else None) for k in _HEADERS},
                "error": err,
            }
        )
        if err:
            errors += 1
        else:
            valid += 1

    job_id = str(uuid.uuid4())
    summary = {"total": len(parsed), "valid": valid, "errors": errors}
    await s.execute(
        text(
            "INSERT INTO import_jobs (id, tenant_id, filename, status, summary, created_by) "
            "VALUES (:id, :t, :fn, 'validated', CAST(:sm AS jsonb), :by)"
        ),
        {"id": job_id, "t": tenant_id, "fn": filename, "sm": json.dumps(summary), "by": created_by},
    )
    for row in parsed:
        await s.execute(
            text(
                "INSERT INTO import_rows (tenant_id, job_id, row_no, data, error, action) "
                "VALUES (:t, :j, :rn, CAST(:d AS jsonb), :e, :a)"
            ),
            {
                "t": tenant_id,
                "j": job_id,
                "rn": row["row_no"],
                "d": json.dumps(row["data"], ensure_ascii=False),
                "e": row["error"],
                "a": "error" if row["error"] else "create",
            },
        )
    return {"job_id": job_id, "summary": summary, "rows": parsed}


async def commit_job(s: AsyncSession, tenant_id: str, job_id: str, creator_role: str) -> list[dict]:
    status = (
        await s.execute(text("SELECT status FROM import_jobs WHERE id = :id"), {"id": job_id})
    ).scalar_one_or_none()
    if status is None:
        raise KeyError("job_not_found")
    if status == "committed":
        raise ValueError("already_committed")

    class_map = await _class_map(s)
    rows = (
        (
            await s.execute(
                text(
                    "SELECT row_no, data FROM import_rows "
                    "WHERE job_id = :j AND action = 'create' ORDER BY row_no"
                ),
                {"j": job_id},
            )
        )
        .mappings()
        .all()
    )

    creds: list[dict] = []
    for r in rows:
        d = r["data"]
        class_id = class_map.get(d.get("class_name")) if d.get("class_name") else None
        cred = await us.create_user(
            s,
            tenant_id,
            creator_role,
            {
                "full_name": d["full_name"],
                "role": "student",
                "dob": d.get("dob"),
                "parent_phone": d.get("parent_phone"),
                "class_id": class_id,
            },
        )
        creds.append(cred)
    await s.execute(
        text("UPDATE import_jobs SET status = 'committed' WHERE id = :id"), {"id": job_id}
    )
    return creds
